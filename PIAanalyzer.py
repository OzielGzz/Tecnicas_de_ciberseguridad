from virus_total_apis import PublicApi
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import argparse

try:
    def GuardarInformacion(response):
        wb = load_workbook('reporte_analizador_urls.xlsx')
        hoja = wb["resultados_obetenidos"]

        val = response.values()
        val = list(val)
        dic = val[0]
        dicval = dic.values()
        dicval = list(dicval)
        ext = dicval
        du = ext[8]
        dv = int(du)

        if dv < 3:
            clas = "Baja"
        elif dv > 3 and dv < 10:
            clas = "Media"
        else:
            clas = "Alta"

        wb.save("reporte_analizador_urls.xlsx")

        hoja = wb["resultados_obetenidos"]
        hoja.append([ext[2], ext[4], ext[9], ext[8], clas])
        wb.save("reporte_analizador_urls.xlsx")

    def inicio(Key, Urls):
        api = PublicApi(Key)

        libro = Workbook()
        hoja = libro.create_sheet("resultados_obetenidos")
        libro.save("reporte_analizador_urls.xlsx")
        hoja = libro["resultados_obetenidos"]
        hoja.append(["URL", "DiaAnalis ", "AnalisTot", "AnálisPos", "Clasific"])
        libro.save("reporte_analizador_urls.xlsx")

        with open(Urls, "r") as archivo:
            for linea in archivo:
                response = api.get_url_report(linea)
                GuardarInformacion(response)
                time.sleep(15)
                print('han pasado 15 segundos')

    if __name__ == '__main__':
        inicio(args.Key, args.Urls)

except SyntaxError:
    print("")
